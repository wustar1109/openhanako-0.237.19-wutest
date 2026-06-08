#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path

from office_common import (
    DocumentSkillError,
    json_error,
    json_success,
    load_ops,
    prepare_output,
    require_file,
    write_text_or_json,
)


def require_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except Exception as exc:
        raise DocumentSkillError(
            "XLSX editing requires the optional dependency openpyxl.",
            code="missing_dependency",
            details={"package": "openpyxl", "original_error": str(exc)},
        ) from exc


def sheet_by_name(workbook, name: str | None):
    if name:
        if name not in workbook.sheetnames:
            raise DocumentSkillError(f"Sheet not found: {name}", code="sheet_not_found")
        return workbook[name]
    return workbook.active


def normalize_color(value):
    if value is None:
        return None
    text = str(value).strip().replace("#", "")
    if len(text) == 6:
        return text.upper()
    if len(text) == 8:
        return text.upper()
    raise DocumentSkillError(f"Invalid color value: {value}", code="invalid_color")


def run_ops(input_path: Path, output_path: Path, ops: list[dict]) -> list[dict]:
    openpyxl = require_openpyxl()
    from openpyxl.styles import Font, PatternFill  # type: ignore

    keep_vba = input_path.suffix.lower() == ".xlsm"
    workbook = openpyxl.load_workbook(str(input_path), keep_vba=keep_vba)
    results: list[dict] = []

    for op in ops:
        kind = op["op"]
        if kind == "set_cell":
            sheet = sheet_by_name(workbook, op.get("sheet"))
            cell = op.get("cell")
            if not isinstance(cell, str):
                raise DocumentSkillError("set_cell requires a cell like A1.", code="invalid_operation")
            sheet[cell] = op.get("value")
            results.append({"op": kind, "sheet": sheet.title, "cell": cell})
        elif kind == "set_formula":
            sheet = sheet_by_name(workbook, op.get("sheet"))
            cell = op.get("cell")
            formula = op.get("formula")
            if not isinstance(cell, str) or not isinstance(formula, str):
                raise DocumentSkillError("set_formula requires cell and formula strings.", code="invalid_operation")
            sheet[cell] = formula if formula.startswith("=") else "=" + formula
            results.append({"op": kind, "sheet": sheet.title, "cell": cell})
        elif kind == "append_row":
            sheet = sheet_by_name(workbook, op.get("sheet"))
            values = op.get("values")
            if not isinstance(values, list):
                raise DocumentSkillError("append_row requires a values list.", code="invalid_operation")
            sheet.append(values)
            results.append({"op": kind, "sheet": sheet.title, "row": sheet.max_row})
        elif kind == "add_sheet":
            name = op.get("name")
            if not isinstance(name, str) or not name:
                raise DocumentSkillError("add_sheet requires a non-empty name.", code="invalid_operation")
            if name in workbook.sheetnames:
                raise DocumentSkillError(f"Sheet already exists: {name}", code="sheet_exists")
            index = op.get("index")
            if index is None:
                workbook.create_sheet(name)
            else:
                workbook.create_sheet(name, int(index))
            results.append({"op": kind, "sheet": name})
        elif kind == "rename_sheet":
            sheet = sheet_by_name(workbook, op.get("sheet"))
            name = op.get("name")
            if not isinstance(name, str) or not name:
                raise DocumentSkillError("rename_sheet requires a non-empty name.", code="invalid_operation")
            if name in workbook.sheetnames and name != sheet.title:
                raise DocumentSkillError(f"Sheet already exists: {name}", code="sheet_exists")
            old = sheet.title
            sheet.title = name
            results.append({"op": kind, "from": old, "to": name})
        elif kind == "delete_sheet":
            sheet = sheet_by_name(workbook, op.get("sheet"))
            if len(workbook.sheetnames) <= 1:
                raise DocumentSkillError("Cannot delete the only sheet in a workbook.", code="invalid_operation")
            title = sheet.title
            workbook.remove(sheet)
            results.append({"op": kind, "sheet": title})
        elif kind == "set_number_format":
            sheet = sheet_by_name(workbook, op.get("sheet"))
            cell = op.get("cell")
            fmt = op.get("format")
            if not isinstance(cell, str) or not isinstance(fmt, str):
                raise DocumentSkillError("set_number_format requires cell and format strings.", code="invalid_operation")
            sheet[cell].number_format = fmt
            results.append({"op": kind, "sheet": sheet.title, "cell": cell})
        elif kind == "set_style":
            sheet = sheet_by_name(workbook, op.get("sheet"))
            cell_ref = op.get("cell")
            if not isinstance(cell_ref, str):
                raise DocumentSkillError("set_style requires a cell string.", code="invalid_operation")
            cell = sheet[cell_ref]
            font_kwargs = {}
            if "bold" in op:
                font_kwargs["bold"] = bool(op["bold"])
            if "italic" in op:
                font_kwargs["italic"] = bool(op["italic"])
            if "font_color" in op:
                font_kwargs["color"] = normalize_color(op["font_color"])
            if font_kwargs:
                cell.font = cell.font.copy(**font_kwargs)
            if "fill_color" in op:
                cell.fill = PatternFill("solid", fgColor=normalize_color(op["fill_color"]))
            results.append({"op": kind, "sheet": sheet.title, "cell": cell_ref})
        elif kind == "autofit":
            sheet = sheet_by_name(workbook, op.get("sheet"))
            max_width = int(op.get("max_width", 60))
            for column_cells in sheet.columns:
                first = column_cells[0]
                column = first.column_letter
                width = min(max_width, max(len(str(cell.value or "")) for cell in column_cells) + 2)
                sheet.column_dimensions[column].width = max(width, 8)
            results.append({"op": kind, "sheet": sheet.title})
        else:
            raise DocumentSkillError(f"Unsupported XLSX operation: {kind}", code="unsupported_operation")

    workbook.save(str(output_path))
    verify_workbook = openpyxl.load_workbook(str(output_path), data_only=False, keep_vba=keep_vba)
    formula_count = 0
    error_literals = []
    for sheet in verify_workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.startswith("="):
                        formula_count += 1
                    if cell.value in {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}:
                        error_literals.append(f"{sheet.title}!{cell.coordinate}:{cell.value}")
    results.append({"op": "verify_reload", "sheets": verify_workbook.sheetnames, "formulas": formula_count, "literal_errors": error_literals})
    return results


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply safe, limited edits to XLSX and XLSM files.")
    parser.add_argument("input")
    parser.add_argument("output")
    parser.add_argument("--ops", required=True, help="JSON operations file.")
    args = parser.parse_args()

    try:
        input_path = require_file(args.input)
        output_path = prepare_output(input_path, args.output)
        if input_path.suffix.lower() not in {".xlsx", ".xlsm"} or output_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise DocumentSkillError("edit_xlsx.py requires .xlsx or .xlsm input and output paths.", code="unsupported_format")
        ops = load_ops(args.ops)
        results = run_ops(input_path, output_path, ops)
        write_text_or_json(None, json_success(input=str(input_path), output=str(output_path), operations=results))
        return 0
    except Exception as exc:
        write_text_or_json(None, json_error(exc))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
